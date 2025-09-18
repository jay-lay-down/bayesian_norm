# ===== 2) 모듈 저장 (adj_limit/whitelist + cores 옵션 추가, bxj 끌 수 있음) =====
module_code = r'''
import os, argparse
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import imageio.v2 as imageio

def clean_txt(s):
    if s is None or (isinstance(s, float) and pd.isna(s)): return ""
    s = str(s).replace("\t", " ").replace("\xa0", " ")
    return " ".join(s.split()).strip()

def key_txt(s):
    t = clean_txt(s).lower()
    for ch in [" ", "_", "-"]: t = t.replace(ch, "")
    return t

def find_col(df, candidates):
    cols = list(df.columns)
    for c in cols:
        for cand in candidates:
            if clean_txt(c).lower() == clean_txt(cand).lower(): return c
    keys = [key_txt(c) for c in cols]
    for i, c in enumerate(cols):
        for cand in candidates:
            if key_txt(cand) == keys[i]: return c
    for c in cols:
        lc = str(c).lower()
        for cand in candidates:
            if str(cand).lower() in lc: return c
    return None

def read_excel_robust(path, sheet_name=0):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx",".xlsm",".xltx",".xltm",""]:
        try: return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        except Exception: pass
    if ext == ".xls":
        try: return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
        except Exception: pass
    if ext == ".xlsb":
        try: return pd.read_excel(path, sheet_name=sheet_name, engine="pyxlsb")
        except Exception: pass
    try: return pd.read_csv(path)
    except Exception: return pd.read_excel(path, sheet_name=sheet_name)

def fit_hb(
    df_std, seed=2025, draws=2000, tune=2000, chains=4,
    target_accept=0.9, max_treedepth=12,
    include_bxj=True, include_bxa=False, include_axj=False,
    use_normal_lik=False, sample_nu=True, cores=None
):
    import pymc as pm
    y   = df_std["y_std"].values.astype("float32")
    b_i = df_std["b_i"].values
    a_i = df_std["a_i"].values
    j_i = df_std["j_i"].values

    brand_ix = {b:i for i,b in enumerate(sorted(df_std["brand"].unique()))}
    attr_ix  = {a:i for i,a in enumerate(sorted(df_std["attr"].unique()))}
    adj_ix   = {w:i for i,w in enumerate(sorted(df_std["adj"].unique()))}

    coords = {"brand": list(brand_ix.keys()),
              "attr":  list(attr_ix.keys()),
              "adj":   list(adj_ix.keys()),
              "obs_id": np.arange(len(y))}

    with pm.Model(coords=coords) as model:
        pm.Data("b_i", b_i, dims="obs_id")
        pm.Data("a_i", a_i, dims="obs_id")
        pm.Data("j_i", j_i, dims="obs_id")

        mu0 = pm.Normal("mu0", 0.0, 1.0)
        sigma_obs = pm.HalfNormal("sigma_obs", 1.0)

        sd_brand = pm.HalfNormal("sd_brand", 1.0)
        sd_attr  = pm.HalfNormal("sd_attr",  1.0)
        sd_adj   = pm.HalfNormal("sd_adj",   1.0)

        z_brand = pm.Normal("z_brand", 0, 1, dims=("brand",))
        z_attr  = pm.Normal("z_attr",  0, 1, dims=("attr",))
        z_adj   = pm.Normal("z_adj",   0, 1, dims=("adj",))

        brand_eff = pm.Deterministic("brand_eff", z_brand * sd_brand, dims=("brand",))
        attr_eff  = pm.Deterministic("attr_eff",  z_attr  * sd_attr,  dims=("attr",))
        adj_eff   = pm.Deterministic("adj_eff",   z_adj   * sd_adj,   dims=("adj",))

        if include_bxj:
            sd_ba = pm.HalfNormal("sd_ba", 0.5)
            z_ba  = pm.Normal("z_ba", 0, 1, dims=("brand","adj"))
            ba_eff = pm.Deterministic("ba_eff", z_ba * sd_ba, dims=("brand","adj"))
        else:
            ba_eff = 0.0
        if include_bxa:
            sd_ba2 = pm.HalfNormal("sd_ba2", 0.5)
            z_ba2  = pm.Normal("z_ba2", 0, 1, dims=("brand","attr"))
            ba2_eff = pm.Deterministic("ba2_eff", z_ba2 * sd_ba2, dims=("brand","attr"))
        else:
            ba2_eff = 0.0
        if include_axj:
            sd_aj = pm.HalfNormal("sd_aj", 0.5)
            z_aj  = pm.Normal("z_aj", 0, 1, dims=("attr","adj"))
            aj_eff = pm.Deterministic("aj_eff", z_aj * sd_aj, dims=("attr","adj"))
        else:
            aj_eff = 0.0

        mu = mu0 + brand_eff[b_i] + attr_eff[a_i] + adj_eff[j_i]
        if include_bxj: mu = mu + ba_eff[b_i, j_i]
        if include_bxa: mu = mu + ba2_eff[b_i, a_i]
        if include_axj: mu = mu + aj_eff[a_i, j_i]

        if use_normal_lik:
            pm.Normal("y", mu=mu, sigma=sigma_obs, observed=y, dims=("obs_id",))
        else:
            if sample_nu:
                nu_raw = pm.Exponential("nu_raw", 1/30)
                nu = pm.Deterministic("nu", nu_raw + 2.0)
                pm.StudentT("y", nu=nu, mu=mu, sigma=sigma_obs, observed=y, dims=("obs_id",))
            else:
                pm.Deterministic("nu", 30.0)
                pm.StudentT("y", nu=30.0, mu=mu, sigma=sigma_obs, observed=y, dims=("obs_id",))

        step = pm.NUTS(target_accept=target_accept, max_treedepth=max_treedepth)
        idata = pm.sample(draws=draws, tune=tune, chains=chains, step=step,
                          random_seed=seed, init="jitter+adapt_diag",
                          cores=cores if cores is not None else 1, progressbar=True)
    return idata, {"brand": list(brand_ix.keys()),
                   "attr":  list(attr_ix.keys()),
                   "adj":   list(adj_ix.keys())}, brand_ix, attr_ix, adj_ix

def tables_from_post(df_long, idata, coords, brand_ix, attr_ix, adj_ix, y_mean, y_std):
    post = idata.posterior
    def samp(name, dims=None):
        x = post[name].stack(s=("chain","draw"))
        if dims is None: return x.transpose("s").values
        return x.transpose("s", *dims).values
    mu0      = samp("mu0")
    sigma    = samp("sigma_obs")
    nu_draw  = samp("nu") if ("nu" in post.variables) else None
    brand_fx = samp("brand_eff", ("brand",))
    attr_fx  = samp("attr_eff",  ("attr",))
    adj_fx   = samp("adj_eff",   ("adj",))
    ba_fx    = samp("ba_eff",  ("brand","adj")) if ("ba_eff"  in post) else None
    ba2_fx   = samp("ba2_eff", ("brand","attr")) if ("ba2_eff" in post) else None
    aj_fx    = samp("aj_eff",  ("attr","adj"))   if ("aj_eff"  in post) else None

    cnt_overall = (df_long.groupby(["brand","attr"], dropna=False)
                   .size().reset_index(name="n"))
    cnt_overall["w"] = cnt_overall["n"] / cnt_overall["n"].sum()

    cnt_brand_attr = (df_long.groupby(["brand","attr"], dropna=False)
                      .size().reset_index(name="n"))
    brand_weights = {}
    attrs = sorted(df_long["attr"].unique().tolist())
    for b, sub in cnt_brand_attr.groupby("brand"):
        sub = sub.copy(); sub["w"] = sub["n"] / sub["n"].sum()
        brand_weights[b] = dict(zip(sub["attr"], sub["w"]))

    brands = list(brand_ix.keys())
    adjs   = list(adj_ix.keys())

    def mu_draws_baj(b, a, j):
        bi = brand_ix[b]; ai = attr_ix[a]; ji = adj_ix[j]
        m = mu0 + brand_fx[:, bi] + attr_fx[:, ai] + adj_fx[:, ji]
        if ba_fx  is not None: m = m + ba_fx[:, bi, ji]
        if ba2_fx is not None: m = m + ba2_fx[:, bi, ai]
        if aj_fx  is not None: m = m + aj_fx[:, ai, ji]
        return m

    def expected_obs_var():
        if nu_draw is None: return float(np.mean(sigma**2))
        return float(np.mean((sigma**2) * (nu_draw / np.maximum(nu_draw-2.0, 1e-6))))
    obs_var = expected_obs_var()

    def pred_stats_overall_for_adj(adj_name):
        vals, weights = [], []
        for _, row in cnt_overall.iterrows():
            b = row["brand"]; a = row["attr"]; w = float(row["w"])
            vals.append(mu_draws_baj(b, a, adj_name)); weights.append(w)
        vals = np.vstack(vals); weights = np.asarray(weights).reshape(-1,1)
        mu_draws = (weights * vals).sum(axis=0)
        var_mu = float(np.var(mu_draws, ddof=0))
        var_pred = var_mu + obs_var
        return float(np.mean(mu_draws)), float(np.sqrt(max(var_pred, 1e-12)))

    def pred_stats_brand_for_adj(brand_name, adj_name):
        w_map = brand_weights.get(brand_name, None) or {a: 1.0/len(attrs) for a in attrs}
        vals, weights = [], []
        for a, w in w_map.items():
            vals.append(mu_draws_baj(brand_name, a, adj_name)); weights.append(float(w))
        vals = np.vstack(vals); weights = np.asarray(weights).reshape(-1,1)
        mu_draws = (weights * vals).sum(axis=0)
        var_mu = float(np.var(mu_draws, ddof=0))
        var_pred = var_mu + obs_var
        return float(np.mean(mu_draws)), float(np.sqrt(max(var_pred, 1e-12)))

    bayes_rows, reg_rows = [], []
    alpha_base = 30.0/2.0 if (nu_draw is None) else float(np.mean(nu_draw))/2.0
    alpha_base = max(alpha_base, 1.01)

    for j in adjs:
        mu_bar_std, sd_pred_std = pred_stats_overall_for_adj(j)
        kappa, alpha = 1.0, alpha_base
        sc2   = (sd_pred_std**2) * (2*alpha - 2.0) / max(2*alpha, 2.000001)
        beta  = sc2 * alpha * kappa / (kappa + 1.0)
        mu_bar = mu_bar_std * y_std + y_mean
        sd_pred = sd_pred_std * y_std
        bayes_rows.append({"Group":"전체","Adjective":j,"post_mu":mu_bar,"post_k":kappa,
                           "post_alpha":alpha,"post_beta":beta})
        reg_rows.append({"Group":"전체","Adjective":j,"mean":mu_bar,"sd":sd_pred})

    for b in brands:
        for j in adjs:
            mu_bar_std, sd_pred_std = pred_stats_brand_for_adj(b, j)
            kappa, alpha = 1.0, alpha_base
            sc2   = (sd_pred_std**2) * (2*alpha - 2.0) / max(2*alpha, 2.000001)
            beta  = sc2 * alpha * kappa / (kappa + 1.0)
            mu_bar = mu_bar_std * y_std + y_mean
            sd_pred = sd_pred_std * y_std
            bayes_rows.append({"Group":b,"Adjective":j,"post_mu":mu_bar,"post_k":kappa,
                               "post_alpha":alpha,"post_beta":beta})
            reg_rows.append({"Group":b,"Adjective":j,"mean":mu_bar,"sd":sd_pred})

    bayesTbl = pd.DataFrame(bayes_rows).sort_values(["Group","Adjective"], kind="stable").reset_index(drop=True)
    regTbl   = pd.DataFrame(reg_rows).sort_values(["Group","Adjective"], kind="stable").reset_index(drop=True)
    adjTbl   = (df_long[["attr","adj"]].drop_duplicates()
                .rename(columns={"attr":"속성","adj":"Adjective"})
                .sort_values(["속성","Adjective"], kind="stable").reset_index(drop=True))
    return bayesTbl, regTbl, adjTbl

def run_all(
    DATA_PATH, seed=2025, draws=2000, tune=2000, chains=4,
    target_accept=0.9, max_treedepth=12,
    include_bxj=True, include_bxa=False, include_axj=False,
    use_normal_lik=False,
    adj_limit=None, adj_whitelist=None, cores=None
):
    if not os.path.exists(DATA_PATH):
        raise FileNotFoundError(f"데이터 파일을 찾을 수 없습니다: {DATA_PATH}")
    base_dir = os.path.dirname(DATA_PATH) or "."
    out_xlsx = os.path.join(base_dir, "hb_results.xlsx")
    out_gif  = os.path.join(base_dir, "hb_posterior_formation.gif")

    df0 = read_excel_robust(DATA_PATH, sheet_name=0)

    brand_col = find_col(df0, ["brand", "브랜드", "group", "그룹"])
    attr_col  = find_col(df0, ["attribute", "속성", "attr"])
    if not all([brand_col, attr_col]):
        raise ValueError(f"브랜드/속성 컬럼을 못 찾음. brand={brand_col}, attr={attr_col}, columns={list(df0.columns)}")

    id_like_keys = {"respid","id","responseid","respondent","case","rowid"}
    id_like = {c for c in df0.columns if key_txt(c) in id_like_keys}
    protect = set([brand_col, attr_col]) | id_like

    # 후보 형용사 자동 탐지
    adj_candidates = []
    for c in df0.columns:
        if c in protect: continue
        coerced = pd.to_numeric(df0[c], errors="coerce")
        if coerced.notna().sum() > 0: adj_candidates.append(c)

    if len(adj_candidates) == 0:
        adj_col = find_col(df0, ["adjective", "형용사", "adj"])
        y_col   = find_col(df0, ["score", "점수", "rating", "value", "y", "point", "측정값"])
        if not all([adj_col, y_col]):
            raise ValueError("형용사 자동탐지 실패: wide 숫자열도 없고 long(adj/y)도 없음.")
        df = df0[[brand_col, attr_col, adj_col, y_col]].copy()
        df.columns = ["brand","attr","adj","y"]
    else:
        tmp = df0[[brand_col, attr_col] + adj_candidates].copy()
        for c in adj_candidates: tmp[c] = pd.to_numeric(tmp[c], errors="coerce")
        melted = tmp.melt(id_vars=[brand_col, attr_col], value_vars=adj_candidates,
                          var_name="adj", value_name="y")
        df = melted.dropna(subset=["y"]).copy()
        df.rename(columns={brand_col:"brand", attr_col:"attr"}, inplace=True)

    for c in ["brand","attr","adj"]: df[c] = df[c].map(clean_txt)
    df = df.dropna(subset=["brand","attr","adj","y"]).reset_index(drop=True)

    # ---- 속도핵: 형용사 제한/화이트리스트 ----
    if adj_whitelist is not None:
        wl = {clean_txt(x) for x in adj_whitelist}
        df = df[df["adj"].isin(wl)]
    if (adj_limit is not None) and (adj_limit > 0):
        keep = (df["adj"].value_counts().head(adj_limit).index)
        df = df[df["adj"].isin(keep)]

    y_mean, y_std = float(df["y"].mean()), float(df["y"].std(ddof=0))
    if y_std == 0.0: y_std = 1.0
    df["y_std"] = (df["y"] - y_mean) / y_std

    brand_ix = {b:i for i,b in enumerate(sorted(df["brand"].unique()))}
    attr_ix  = {a:i for i,a in enumerate(sorted(df["attr"].unique()))}
    adj_ix   = {w:i for i,w in enumerate(sorted(df["adj"].unique()))}
    df["b_i"] = df["brand"].map(brand_ix).astype("int64")
    df["a_i"] = df["attr"].map(attr_ix).astype("int64")
    df["j_i"] = df["adj"].map(adj_ix).astype("int64")

    print(f"[INFO] Wide→long 변환: rows={len(df)}, brands={len(brand_ix)}, attrs={len(attr_ix)}, adjs={len(adj_ix)}")

    # 샘플링
    idata, coords, brand_ix, attr_ix, adj_ix = fit_hb(
        df, seed=seed, draws=draws, tune=tune, chains=chains,
        target_accept=target_accept, max_treedepth=max_treedepth,
        include_bxj=include_bxj, include_bxa=include_bxa, include_axj=include_axj,
        use_normal_lik=use_normal_lik, sample_nu=not use_normal_lik,
        cores=cores if cores is not None else 2
    )

    bayesTbl, regTbl, adjTbl = tables_from_post(df, idata, coords, brand_ix, attr_ix, adj_ix, y_mean, y_std)

    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    with pd.ExcelWriter(os.path.join(base_dir, "hb_results.xlsx"), engine="openpyxl") as writer:
        bayesTbl.to_excel(writer, index=False, sheet_name="bayesTbl")
        regTbl.to_excel(writer,   index=False, sheet_name="regTbl")
        adjTbl.to_excel(writer,   index=False, sheet_name="adjTbl")
    wb = load_workbook(os.path.join(base_dir, "hb_results.xlsx"))
    def add_table(ws, name):
        from openpyxl.utils import get_column_letter
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        try:
            t = Table(displayName=name, ref=ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(t)
        except Exception:
            pass
    add_table(wb["bayesTbl"], "bayesTbl")
    add_table(wb["regTbl"],   "regTbl")
    add_table(wb["adjTbl"],   "adjTbl")
    wb.save(os.path.join(base_dir, "hb_results.xlsx"))

    print(f"[OK] Saved tables to: {os.path.join(base_dir, 'hb_results.xlsx')}")
    return os.path.join(base_dir, "hb_results.xlsx"), None
'''
mod_path = os.path.join(WORK_DIR, "hb_brand_attr_adj.py")
with open(mod_path, "w", encoding="utf-8") as f:
    f.write(module_code)
print("saved module:", mod_path)
